from random_number_operator import num_list
import openpyxl
import json
import os
import shutil

#############
# file's path
excel_file = '/Users/jhban/Desktop/json/name.xlsx'
image_path = '/Users/jhban/Desktop/build/images/'
image_path_destination = '/Users/jhban/Desktop/n_images/'
json_path = '/Users/jhban/Desktop/build/json/'
json_path_destination = '/Users/jhban/Desktop/n_json/'
tmp_path = '/Users/jhban/Desktop/tmp/'
#############

# open excel file's first sheet
wb = openpyxl.load_workbook(excel_file, read_only=True)
sheet = wb.worksheets[0]

# # it works as a key
# key_list = []
# for col_num in range(1, sheet.max_column):
#     key_list.append(sheet.cell(row=1, column=col_num).value)

#
data_list = []
for row_num in range(1, 510):
    tmp_list = []
    for col_num in range(2, sheet.max_column):
        val = sheet.cell(row=row_num, column=col_num).value
        tmp_list.append(val)
    data_list.append(tmp_list)

wb.close()

for i in range(1, 501):
    with open(os.path.join(json_path, str(i) + '.json'), 'r') as fp:
        data = json.load(fp)

    del data["character"]
    del data["dna"]
    del data["date"]
    del data["compiler"]
    del data["edition"]

    data["name"] = "0xG #" + str(num_list[i])
    data["description"] = data_list[i][0] + ', ' + data_list[i][1]
    data["image"] = "ipfs://BaseUri/" + str(num_list[i]) + ".png"
    # data["edition"] = num_list[i]
    with open(os.path.join(json_path, str(i) + '.json'), 'w', encoding='utf-8') as fp:
        json.dump(data, fp, indent=4)
    shutil.move(json_path + str(i) + ".json", tmp_path)
    shutil.move(image_path + str(i) + ".png", tmp_path)
    os.rename(tmp_path + str(i) + ".json",
              tmp_path + str(num_list[i]) + ".json")
    os.rename(tmp_path + str(i) + ".png",
              tmp_path + str(num_list[i]) + ".png")
    shutil.move(tmp_path + str(num_list[i]) + ".json", json_path_destination)
    shutil.move(tmp_path + str(num_list[i]) + ".png", image_path_destination)
